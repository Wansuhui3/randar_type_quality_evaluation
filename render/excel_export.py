"""Excel report generation with openpyxl - professional styling."""
import io

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from core.models import EvalReport, TrajResult

TYPE_NAMES = {1: 'Car', 2: 'Truck', 3: 'Cyclist', 4: 'Pedestrian', 7: 'Uncertain'}

# === Style constants ===
HEADER_FONT = Font(bold=True, size=10, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)

DATA_FONT = Font(size=9)
DATA_ALIGN_LEFT = Alignment(horizontal='left', vertical='center')
DATA_ALIGN_CENTER = Alignment(horizontal='center', vertical='center')

ROW_FILL_EVEN = PatternFill(start_color='F2F7FC', end_color='F2F7FC', fill_type='solid')
ROW_FILL_ODD = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

FAIL_FONT = Font(size=9, bold=True, color='C00000')
PASS_FONT = Font(size=9, color='2E7D32')

THIN_BORDER = Border(
    left=Side(style='thin', color='B4B4B4'),
    right=Side(style='thin', color='B4B4B4'),
    top=Side(style='thin', color='B4B4B4'),
    bottom=Side(style='thin', color='B4B4B4'),
)

SECTION_FONT = Font(bold=True, size=10, color='2F5496')


def _style_header_row(ws, num_cols, row=1):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 30


def _style_data_cell(ws, row, col, center=False):
    cell = ws.cell(row=row, column=col)
    cell.font = DATA_FONT
    cell.alignment = DATA_ALIGN_CENTER if center else DATA_ALIGN_LEFT
    cell.border = THIN_BORDER
    cell.fill = ROW_FILL_EVEN if row % 2 == 0 else ROW_FILL_ODD
    return cell


def export_report(report: EvalReport, curve_images: dict = None,
                  embed: bool = True) -> bytes:
    wb = Workbook()
    _write_summary(wb, report)
    _write_detail_sheets(wb, report, curve_images, embed)
    _write_field_guide(wb)
    _write_config(wb, report.config)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _write_summary(wb: Workbook, report: EvalReport):
    ws = wb.active
    ws.title = "Summary"

    summary_df = report.summary_df()
    if summary_df.empty:
        ws.cell(row=1, column=1, value="No data")
        return

    headers = [
        'Radar\n雷达来源', 'Type\n类型编号', 'Name\n名称',
        'Trajectories\n航迹数', 'Pass\n通过数', 'Fail\n失败数',
        'Fail Rate\n跳变航迹占比',
        'C-Type Fail\n实测跳变fail', 'Context-ABD Fail\n状态切换fail',
        'Avg Transitions\n平均跳变帧数', 'Avg Duration(s)\n平均时长'
    ]
    df_keys = [
        'Radar', 'Type', 'Name', 'Trajectories', 'Pass', 'Fail',
        'Fail Rate', 'C-Type Fail', 'Context-ABD Fail',
        'Avg Transitions', 'Avg Duration(s)'
    ]

    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, len(headers))

    center_cols = {1, 2, 4, 5, 6, 7, 8, 9, 10, 11}
    for row_idx, (_, row) in enumerate(summary_df.iterrows(), 2):
        for col_idx, key in enumerate(df_keys, 1):
            cell = _style_data_cell(ws, row_idx, col_idx, center=(col_idx in center_cols))
            cell.value = row.get(key, '')

    # Filter log
    log_start = len(summary_df) + 4
    ws.cell(row=log_start, column=1, value="Filter Log 过滤日志").font = SECTION_FONT
    log_headers = ['Step 步骤', 'Input 输入', 'Output 输出', 'Removed 剔除']
    for col, h in enumerate(log_headers, 1):
        ws.cell(row=log_start + 1, column=col, value=h)
    _style_header_row(ws, 4, row=log_start + 1)

    for i, log in enumerate(report.filter_logs):
        r = log_start + 2 + i
        for col, val in enumerate([log.step_name, log.input_count, log.output_count, log.removed_count], 1):
            cell = _style_data_cell(ws, r, col, center=(col > 1))
            cell.value = val

    # Freeze + filter
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(summary_df)+1}"

    summary_widths = [14, 10, 12, 14, 10, 10, 16, 18, 20, 18, 16]
    for i, w in enumerate(summary_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_detail_sheets(wb: Workbook, report: EvalReport,
                         curve_images: dict, embed: bool):
    type_order = [1, 2, 3, 4, 7]
    sheet_names = {
        1: 'Type1_Car', 2: 'Type2_Truck', 3: 'Type3_Cyclist',
        4: 'Type4_Ped', 7: 'Type7_Uncertain'
    }

    headers = [
        'Traj ID\n航迹编号', 'Radar\n雷达来源', 'Start Time\n起始时间',
        'End Time\n结束时间', 'Duration(s)\n持续时长', 'Frames\n帧数',
        'Dominant Type\n主类', 'Transitions\n跳变次数',
        'Transitions(In)\n有效距离内跳变', 'Context\n跳变上下文',
        'ExistProb@Jump\n跳变帧存在概率',
        'MSR Oscillation\n测量振荡', 'MSR Osc Ratio\n振荡占比',
        'Speed(km/h)\n均速', 'Verdict\n判定',
    ]
    if embed and curve_images:
        headers.append('Type Curve\n类型曲线')

    # Columns that should be center-aligned
    center_cols = {2, 5, 6, 7, 8, 9, 10, 12, 13, 14, 15}

    for dtype in type_order:
        type_results = [r for r in report.results if r.dominant_type == dtype]
        if not type_results:
            continue

        ws = wb.create_sheet(title=sheet_names[dtype])

        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        _style_header_row(ws, len(headers))

        # Sort: radar group, fail first (transition_count desc), then pass
        radar_order = {'flr': 0, 'fr': 1, 'rlr': 2, 'rr': 3, 'unk': 4}
        type_results.sort(key=lambda r: (
            radar_order.get(r.radar_src, 9),
            0 if r.verdict == 'fail' else 1,
            -r.transition_count_in,
        ))

        for row_idx, result in enumerate(type_results, 2):
            values = [
                result.traj_id, result.radar_src,
                str(result.start_time), str(result.end_time),
                round(result.duration_s, 1), result.frame_count,
                result.dominant_type, result.transition_count_all,
                result.transition_count_in, ''.join(result.contexts),
                ', '.join(str(e) for e in result.exist_prob_at_jump),
                result.has_msr_oscillation, round(result.msr_osc_ratio, 3),
                round(result.mean_speed_kmh, 1), result.verdict,
            ]
            for col_idx, val in enumerate(values, 1):
                cell = _style_data_cell(ws, row_idx, col_idx, center=(col_idx in center_cols))
                cell.value = val

            # Verdict color
            verdict_cell = ws.cell(row=row_idx, column=15)
            if result.verdict == 'fail':
                verdict_cell.font = FAIL_FONT
            else:
                verdict_cell.font = PASS_FONT

            # Embed curve image
            if embed and curve_images and result.traj_id in curve_images:
                img_bytes = curve_images[result.traj_id]
                img = XlImage(io.BytesIO(img_bytes))
                img.width = 450
                img.height = 175
                col_letter = get_column_letter(16)
                ws.add_image(img, f'{col_letter}{row_idx}')
                ws.row_dimensions[row_idx].height = 135

        # Freeze + filter
        ws.freeze_panes = 'A2'
        last_col = get_column_letter(len(headers))
        ws.auto_filter.ref = f"A1:{last_col}{len(type_results)+1}"

        # Column widths
        col_widths = [18, 8, 20, 20, 11, 8, 10, 11, 13, 12, 20, 12, 10, 10, 8]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        if embed and curve_images:
            ws.column_dimensions[get_column_letter(16)].width = 65


def _write_field_guide(wb: Workbook):
    ws = wb.create_sheet(title="Field Guide")

    guide = [
        ('Traj ID',
         '航迹全局唯一编号，格式为"雷达来源_原始ID_seg序号"（如 flr_12_seg0）。雷达前缀解决跨雷达ID冲突；seg序号由七规则切分算法生成。',
         'Globally unique trajectory ID. Format: {radar}_{original_ID}_seg{N}. Radar prefix resolves cross-radar ID collision; seg index from 7-rule segmentation.'),
        ('Radar',
         '雷达位置标识，从CSV文件名解析：flr=前雷达, rlr=后雷达, fr=前右, rr=后右。若解析失败标记为unk并用Dx符号兜底校验。',
         'Radar position identifier parsed from CSV filename: flr=front, rlr=rear, fr=front-right, rr=rear-right. Falls back to Dx sign check if parsing fails.'),
        ('Context',
         '跳变发生时的 Measured 环境代码。A=进入预测态时跳变；B=重新捕获时跳变；C=连续实测下跳变（最严重）；D=预测态内漂移。多个跳变则拼接显示，如"ACD"。',
         'Measured-state context at each Type transition. A=entering prediction; B=re-acquisition; C=continuous measurement (most severe); D=prediction drift. Concatenated for multiple transitions.'),
        ('ExistProb@Jump',
         '每次跳变帧的 ExistProb（存在概率）值，多个跳变用逗号分隔（如"100, 52"）。高值（≥90）说明跳变发生在正常实测环境（纯逻辑问题）；低值（<70）说明跳变前目标已长时间未被实测更新。仅辅助人工核查，不影响 pass/fail 判定。',
         'ExistProb value at each Type transition frame, comma-separated (e.g. "100, 52"). High(>=90)=logic issue; Low(<70)=prolonged no-measurement. Auxiliary only, does NOT affect pass/fail.'),
        ('MSR Oscillation',
         '是否存在 Measured 状态振荡。判定依据：Measured 字段在 2 秒滑窗内发生 >= 3 次 1↔2 切换。注意：这是测量状态振荡的事实标记，不是"遮挡"的因果判定。',
         'Whether Measured state oscillation is detected. Criteria: Measured flips (1↔2) >= 3 times within 2s window. NOTE: factual measurement-state label, NOT a causal "occlusion" diagnosis.'),
        ('MSR Osc Ratio',
         '被标记为振荡的帧数占航迹总帧数的比例。0=无振荡，1=全程振荡。',
         'Fraction of frames in oscillation-marked intervals. 0=none, 1=entire trajectory.'),
        ('Verdict',
         '最终判定。pass=有效评估距离内 Type 无任何跳变（输出一致）；fail=有效距离内存在至少 1 次 Type 跳变（输出不一致）。有效距离外的跳变记录但不影响判定。跳变到 Uncertain(7) 同样判 fail。',
         'Final verdict. pass=no Type transition within effective eval range; fail=at least 1 transition within range. Out-of-range transitions recorded but do not affect verdict.'),
        ('Dominant Type',
         '航迹中出现帧数最多的 Type 值，用于分类统计归属。注意：主类归属会掩盖跳变方向，需结合类型曲线图查看具体跳变路径。',
         'Most frequent Type value, used for grouping. NOTE: hides transition direction; refer to Type Curve plot for specific jump paths.'),
        ('Transitions',
         '相邻帧 Type 值不同的总次数（含有效距离外的跳变）。例如 Type 序列 [1,1,1,7,1,1] 有 2 次跳变（1→7 和 7→1）。',
         'Total number of adjacent-frame Type changes (including out-of-range). E.g. [1,1,1,7,1,1] has 2 transitions.'),
        ('Transitions(In)',
         '有效评估距离内的跳变次数，是 pass/fail 的判定依据。0=pass，>0=fail。与 Transitions 的差值即为有效距离外的跳变数（不计入判定）。',
         'Number of Type transitions within effective eval range. Basis for pass/fail. 0=pass, >0=fail.'),
    ]

    ws.cell(row=1, column=1, value='Field')
    ws.cell(row=1, column=2, value='中文解释')
    ws.cell(row=1, column=3, value='English Description')
    _style_header_row(ws, 3)

    wrap_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
    for i, (field_name, cn_desc, en_desc) in enumerate(guide, 2):
        c1 = _style_data_cell(ws, i, 1)
        c1.value = field_name
        c1.font = Font(size=9, bold=True)
        c2 = _style_data_cell(ws, i, 2)
        c2.value = cn_desc
        c2.alignment = wrap_align
        c3 = _style_data_cell(ws, i, 3)
        c3.value = en_desc
        c3.alignment = wrap_align
        ws.row_dimensions[i].height = 45

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 55
    ws.column_dimensions['C'].width = 55
    ws.freeze_panes = 'A2'


def _write_config(wb: Workbook, config: dict):
    ws = wb.create_sheet(title="Config")
    ws.cell(row=1, column=1, value='Parameter 参数')
    ws.cell(row=1, column=2, value='Value 值')
    _style_header_row(ws, 2)

    row = 2
    for section, values in config.items():
        cell = _style_data_cell(ws, row, 1)
        cell.value = f"[{section}]"
        cell.font = SECTION_FONT
        row += 1
        if isinstance(values, dict):
            for k, v in values.items():
                c1 = _style_data_cell(ws, row, 1)
                c1.value = f"  {k}"
                c2 = _style_data_cell(ws, row, 2, center=True)
                c2.value = str(v)
                row += 1
        else:
            c2 = _style_data_cell(ws, row, 2, center=True)
            c2.value = str(values)
            row += 1

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 40
    ws.freeze_panes = 'A2'
